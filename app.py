import os
import re
import csv
import uuid
import base64
import logging
import threading
from datetime import datetime, date
from functools import wraps
from io import StringIO, BytesIO

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv, set_key
from openpyxl import load_workbook

from database import get_db, init_db, SVPCET_DEPARTMENTS
from email_service import (send_registration_email, send_vote_confirmation_email,
                            send_election_announcement, send_election_results,
                            send_care_response_email)
from helpers import fetch_location
from face_service import extract_face_encoding, encoding_to_b64, verify_face

# ─────────────────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────────────────
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ieis-super-secret-dev-key-change-me')

# ── Admin credentials ──
app.config['ADMIN_USER'] = 'admin'
app.config['ADMIN_PASS'] = 'admin123'

# ── Care Staff credentials ──
app.config['CARE_USER'] = 'care'
app.config['CARE_PASS'] = 'care123'

# ── Upload folders ──
SYMBOL_UPLOAD_DIR    = os.path.join(app.root_path, 'static', 'uploads', 'symbols')
COMPLAINT_UPLOAD_DIR = os.path.join(app.root_path, 'static', 'uploads', 'complaints')
VOTER_PHOTO_DIR      = os.path.join(app.root_path, 'static', 'uploads', 'voter_photos')

os.makedirs(SYMBOL_UPLOAD_DIR,    exist_ok=True)
os.makedirs(COMPLAINT_UPLOAD_DIR, exist_ok=True)
os.makedirs(VOTER_PHOTO_DIR,      exist_ok=True)

ALLOWED_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp'}
ALLOWED_DOC_EXTS   = {'.jpg', '.jpeg', '.png', '.pdf'}
ALLOWED_EXCEL_EXTS = {'.xlsx', '.xls'}

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# SVPCET roll number format: 22G01A4321
ROLL_NO_REGEX = re.compile(r'^\d{2}G01A\d{4}$', re.IGNORECASE)

# mail stub — email now sent via Resend HTTP API in email_service.py
mail = None


# ─────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────
def election_status(start_date_str, end_date_str):
    """Return 'upcoming', 'active', or 'ended'."""
    today = date.today()
    start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()
    if today < start:
        return 'upcoming'
    elif start <= today <= end:
        return 'active'
    else:
        return 'ended'


def log_admin_action(action, target=None):
    db = get_db()
    db.execute(
        "INSERT INTO admin_log(action, target, timestamp) VALUES(?,?,?)",
        (action, target, datetime.utcnow().isoformat())
    )
    db.commit()
    db.close()


def is_voter_eligible(db, voter, election):
    """Check if voter is eligible for the given election row."""
    etype = election['eligible_type'] or 'department'
    if etype == 'all':
        return True
    elif etype == 'department':
        row = db.execute(
            "SELECT 1 FROM election_eligible_depts WHERE election_id=? AND department=?",
            (election['election_id'], voter['department'])
        ).fetchone()
        return row is not None
    else:  # specific rolls
        row = db.execute(
            "SELECT 1 FROM election_eligible_rolls WHERE election_id=? AND roll_number=?",
            (election['election_id'], voter['roll_number'])
        ).fetchone()
        return row is not None


def get_eligible_voter_emails(db, election_id):
    """Return list of (email, name) tuples for all eligible registered voters."""
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()
    if not election:
        return []

    etype = election['eligible_type'] or 'department'
    if etype == 'all':
        voters = db.execute(
            "SELECT name, email FROM voters WHERE is_registered=1 AND email IS NOT NULL"
        ).fetchall()
    elif etype == 'department':
        depts = [r['department'] for r in db.execute(
            "SELECT department FROM election_eligible_depts WHERE election_id=?", (election_id,)
        ).fetchall()]
        if not depts:
            return []
        placeholders = ','.join('?' * len(depts))
        voters = db.execute(
            f"SELECT name, email FROM voters WHERE department IN ({placeholders}) AND is_registered=1 AND email IS NOT NULL",
            depts
        ).fetchall()
    else:
        rolls = [r['roll_number'] for r in db.execute(
            "SELECT roll_number FROM election_eligible_rolls WHERE election_id=?", (election_id,)
        ).fetchall()]
        if not rolls:
            return []
        placeholders = ','.join('?' * len(rolls))
        voters = db.execute(
            f"SELECT name, email FROM voters WHERE roll_number IN ({placeholders}) AND is_registered=1 AND email IS NOT NULL",
            rolls
        ).fetchall()

    return [(v['email'], v['name']) for v in voters]


def parse_rolls_from_excel(file_stream):
    """Extract roll numbers from an uploaded .xlsx file. Returns list of strings."""
    rolls = []
    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell:
                    val = str(cell).strip().upper()
                    if ROLL_NO_REGEX.match(val):
                        rolls.append(val)
    except Exception as e:
        logger.error(f"Excel parse error: {e}")
    return rolls


def save_base64_photo(b64_photo, roll_number):
    """Save base64 photo to disk, return db path or None."""
    try:
        header, encoded = b64_photo.split(',', 1)
        ext = '.jpg' if ('jpeg' in header or 'jpg' in header) else '.png'
        file_data = base64.b64decode(encoded)
        filename = f"{roll_number}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = os.path.join(VOTER_PHOTO_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(file_data)
        return f"/static/uploads/voter_photos/{filename}"
    except Exception as e:
        logger.error(f"Photo save failed: {e}")
        return None


# ─────────────────────────────────────────────────────────
# DECORATORS
# ─────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


def care_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('care_logged_in'):
            return redirect(url_for('care_login'))
        return f(*args, **kwargs)
    return decorated


def voter_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('voter_id'):
            return redirect(url_for('voter_login'))
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────────────────
# PUBLIC ROUTES
# ─────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# ─────────────────────────────────────────────────────────
# VOTER REGISTRATION
# ─────────────────────────────────────────────────────────
@app.route('/register', methods=['GET'])
def register():
    return render_template('voter/register.html', departments=SVPCET_DEPARTMENTS)


@app.route('/api/verify_voter', methods=['POST'])
def api_verify_voter():
    data = request.get_json(silent=True) or {}
    roll_number = (data.get('roll_number') or '').strip().upper()
    name        = (data.get('name')        or '').strip()
    email       = (data.get('email')       or '').strip().lower()
    phone       = (data.get('phone')       or '').strip()
    department  = (data.get('department')  or '').strip()

    if not ROLL_NO_REGEX.match(roll_number):
        return jsonify(error='Invalid roll number format.'), 400
    if not re.match(r'^[A-Za-z ]+$', name):
        return jsonify(error='Name must contain alphabets only'), 400
    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return jsonify(error='Please enter a valid email address'), 400
    if not re.match(r'^\d{10}$', phone):
        return jsonify(error='Phone must be exactly 10 digits'), 400
    if not department:
        return jsonify(error='Please select your department'), 400

    db = get_db()
    voter = db.execute("SELECT * FROM voters WHERE roll_number=?", (roll_number,)).fetchone()
    
    if not voter:
        db.close()
        return jsonify(error='Roll number not recognized. Please contact customer care.'), 404
        
    if voter['name'].strip().lower() != name.lower():
        db.close()
        return jsonify(error='Name does not match our records. Check spelling or contact support.'), 400
        
    if voter['is_registered'] == 1:
        db.close()
        return jsonify(error='This roll number is already registered. Please sign in.'), 409
        
    db.close()
    return jsonify(success=True), 200


@app.route('/register', methods=['POST'])
def register_post():
    data = request.get_json(silent=True) or {}
    roll_number = (data.get('roll_number') or '').strip().upper()
    name        = (data.get('name')        or '').strip()
    email       = (data.get('email')       or '').strip().lower()
    phone       = (data.get('phone')       or '').strip()
    photo       = (data.get('photo')       or '').strip()
    password    = (data.get('password')    or '').strip()
    department  = (data.get('department')  or '').strip()

    # ── Server-side validation ──
    if not ROLL_NO_REGEX.match(roll_number):
        return jsonify(error='Invalid roll number format. Expected format: 22G01A4321', code='INVALID_ROLL'), 400
    if not re.match(r'^[A-Za-z ]+$', name):
        return jsonify(error='Name must contain alphabets only', code='INVALID_NAME'), 400
    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return jsonify(error='Please enter a valid email address', code='INVALID_EMAIL'), 400
    if not re.match(r'^\d{10}$', phone):
        return jsonify(error='Phone must be exactly 10 digits', code='INVALID_PHONE'), 400
    if len(photo) < 100:
        return jsonify(error='Face photo is required. Please allow camera access.', code='MISSING_PHOTO'), 400
    if len(password) < 6:
        return jsonify(error='Password must be at least 6 characters', code='INVALID_PASS'), 400
    if not department:
        return jsonify(error='Please select your department', code='MISSING_DEPT'), 400

    # ── DB checks ──
    db = get_db()
    voter = db.execute(
        "SELECT * FROM voters WHERE roll_number=?", (roll_number,)
    ).fetchone()

    if not voter:
        db.close()
        return jsonify(error='Roll number not registered in our system. Contact customer care.', code='NOT_FOUND'), 404

    if voter['name'].strip().lower() != name.lower():
        db.close()
        return jsonify(error='Name does not match our records. Please contact customer care.', code='NAME_MISMATCH'), 400

    if voter['is_registered'] == 1:
        db.close()
        return jsonify(error='This roll number is already registered.', code='DUPLICATE'), 409

    # ── Save photo ──
    db_photo_path = save_base64_photo(photo, roll_number)
    if not db_photo_path:
        db.close()
        return jsonify(error='Failed to process photo. Please try again.', code='PHOTO_ERROR'), 400

    # ── Hash password ──
    pw_hash = generate_password_hash(password)

    # ── Save registration with empty face data (encoded in background) ──
    db.execute(
        "UPDATE voters SET email=?, phone=?, photo=?, face_encoding=NULL, password_hash=?, department=?, is_registered=1 WHERE voter_id=?",
        (email, phone, db_photo_path, pw_hash, department, voter['voter_id'])
    )
    db.commit()

    voter_id = voter['voter_id']
    voter_name = voter['name']
    db.close()

    logger.info(f"DB save done for {roll_number}. Starting background thread...")

    # ── Encode face + send email in background thread ──
    def _encode_face_background(vid, photo_b64, v_email, v_name, v_roll):
        with app.app_context():
            try:
                import time
                time.sleep(1)  # Let HTTP request finish first
    
                # Send registration email (moved here so it doesn't block HTTP response)
                try:
                    from email_service import send_registration_email
                    send_registration_email(mail, v_email, v_name, v_roll)
                    logger.info(f"Registration email sent for {v_roll}")
                except Exception as e:
                    logger.error(f"Registration email error: {e}")
    
                # Encode face via Hugging Face API
                from face_service import extract_face_encoding, encoding_to_b64
                logger.info(f"Starting face encoding for voter {vid} via HF API...")
                embedding = extract_face_encoding(photo_b64)
                if embedding is not None:
                    enc_b64 = encoding_to_b64(embedding)
                    bg_db = get_db()
                    bg_db.execute("UPDATE voters SET face_encoding=? WHERE voter_id=?", (enc_b64, vid))
                    bg_db.commit()
                    bg_db.close()
                    logger.info(f"Face encoded and saved for voter {vid}")
                else:
                    logger.warning(f"Face extraction returned None for voter {vid}")
            except Exception as ex:
                logger.error(f"Background thread exception for voter {vid}: {ex}")

    import threading
    t = threading.Thread(target=_encode_face_background, args=(voter_id, photo, email, voter_name, roll_number), daemon=True)
    t.start()

    logger.info(f"Returning 200 success for {roll_number}")
    return jsonify(success=True, redirect='/voter/login'), 200


# ─────────────────────────────────────────────────────────
# VOTER AUTH
# ─────────────────────────────────────────────────────────
@app.route('/voter/login', methods=['GET', 'POST'])
def voter_login():
    if request.method == 'POST':
        roll_number = request.form.get('roll_number', '').strip().upper()
        password    = request.form.get('password', '').strip()
        db = get_db()
        voter = db.execute(
            "SELECT * FROM voters WHERE roll_number=? AND is_registered=1",
            (roll_number,)
        ).fetchone()
        db.close()

        if not voter:
            flash('Roll number not found or not yet registered.', 'error')
            return redirect(url_for('voter_login'))

        # Check password (allow login if no password_hash set yet — backward compat)
        if voter['password_hash']:
            if not check_password_hash(voter['password_hash'], password):
                flash('Incorrect password.', 'error')
                return redirect(url_for('voter_login'))
        else:
            # Legacy accounts without password — let them in but warn
            flash('Please re-register to set a password for your account.', 'warning')

        session['voter_id']   = voter['voter_id']
        session['voter_name'] = voter['name']
        session['voter_roll'] = voter['roll_number']
        return redirect(url_for('voter_dashboard'))
    return render_template('voter/login.html')


@app.route('/voter/logout')
def voter_logout():
    session.pop('voter_id',   None)
    session.pop('voter_name', None)
    session.pop('voter_roll', None)
    return redirect(url_for('voter_login'))


@app.route('/voter/relink-face', methods=['POST'])
@voter_required
def voter_relink_face():
    """
    Allows a voter to re-submit their face photo for encoding when the
    background thread during registration failed (HF Space cold start).
    This runs the encoding synchronously so we can tell the user if it worked.
    """
    data = request.get_json(silent=True) or {}
    face_photo = data.get('face_photo', '')
    if not face_photo or len(face_photo) < 100:
        return jsonify(success=False, error='No face photo received.'), 400

    voter_id = session['voter_id']
    db = get_db()
    voter = db.execute("SELECT * FROM voters WHERE voter_id=?", (voter_id,)).fetchone()
    if not voter:
        db.close()
        return jsonify(success=False, error='Voter not found.'), 404

    if voter['face_encoding']:
        db.close()
        return jsonify(success=True, message='Face profile already exists.'), 200

    # Run encoding synchronously so the user gets immediate feedback
    embedding = extract_face_encoding(face_photo)
    if embedding is None:
        db.close()
        return jsonify(
            success=False,
            error='Face processing server is still waking up. Please wait 30 seconds and try again.'
        ), 503

    enc_b64 = encoding_to_b64(embedding)
    db.execute("UPDATE voters SET face_encoding=? WHERE voter_id=?", (enc_b64, voter_id))
    db.commit()
    db.close()
    logger.info(f"Face re-linked successfully for voter {voter_id}")
    return jsonify(success=True, message='Face profile saved! You can now vote.'), 200




# ─────────────────────────────────────────────────────────
# VOTER DASHBOARD
# ─────────────────────────────────────────────────────────
@app.route('/voter/dashboard')
@voter_required
def voter_dashboard():
    db = get_db()
    voter = db.execute(
        "SELECT * FROM voters WHERE voter_id=?",
        (session['voter_id'],)
    ).fetchone()

    elections_raw = db.execute("SELECT * FROM elections ORDER BY election_id DESC").fetchall()
    elections = []
    for e in elections_raw:
        status       = election_status(e['start_date'], e['end_date'])
        eligible     = is_voter_eligible(db, voter, e)
        already_voted = bool(db.execute(
            "SELECT 1 FROM votes WHERE voter_id=? AND election_id=?",
            (voter['voter_id'], e['election_id'])
        ).fetchone())
        elections.append({
            'election_id':        e['election_id'],
            'election_title':     e['election_title'],
            'description':        e['description'] or '',
            'position_role':      e['position_role'] or '',
            'start_date':         e['start_date'],
            'end_date':           e['end_date'],
            'status':             status,
            'eligible':           eligible,
            'already_voted':      already_voted,
            'results_published':  e['results_published'],
        })
    db.close()
    return render_template('voter/dashboard.html', voter=voter, elections=elections)


# ─────────────────────────────────────────────────────────
# VOTING
# ─────────────────────────────────────────────────────────
@app.route('/voter/vote/<int:election_id>', methods=['GET'])
@voter_required
def vote_page(election_id):
    db = get_db()
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()

    if not election:
        db.close()
        abort(404)

    voter = db.execute(
        "SELECT * FROM voters WHERE voter_id=?", (session['voter_id'],)
    ).fetchone()

    status = election_status(election['start_date'], election['end_date'])
    if status != 'active':
        flash('This election is not currently active.', 'error')
        db.close()
        return redirect(url_for('voter_dashboard'))

    if not is_voter_eligible(db, voter, election):
        flash('You are not eligible for this election.', 'error')
        db.close()
        return redirect(url_for('voter_dashboard'))

    already_voted = db.execute(
        "SELECT 1 FROM votes WHERE voter_id=? AND election_id=?",
        (voter['voter_id'], election_id)
    ).fetchone()
    if already_voted:
        flash('You have already voted in this election.', 'error')
        db.close()
        return redirect(url_for('voter_dashboard'))

    candidates = db.execute(
        "SELECT * FROM candidates WHERE election_id=?", (election_id,)
    ).fetchall()
    db.close()
    return render_template('voter/vote.html', election=election, candidates=candidates)


@app.route('/voter/vote/<int:election_id>', methods=['POST'])
@voter_required
def vote_submit(election_id):
    data         = request.get_json(silent=True) or {}
    candidate_id = data.get('candidate_id')
    face_photo   = data.get('face_photo', '')  # base64 live capture

    if not candidate_id:
        return jsonify(error='No candidate selected', code='NO_CANDIDATE'), 400

    db = get_db()
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()
    voter = db.execute(
        "SELECT * FROM voters WHERE voter_id=?", (session['voter_id'],)
    ).fetchone()

    if not election:
        db.close()
        return jsonify(error='Election not found', code='NOT_FOUND'), 404

    # Re-run ALL server-side checks
    status = election_status(election['start_date'], election['end_date'])
    if status != 'active':
        db.close()
        return jsonify(error='Election is not active', code='NOT_ACTIVE'), 403

    if not is_voter_eligible(db, voter, election):
        db.close()
        return jsonify(error='You are not eligible for this election', code='NOT_ELIGIBLE'), 403

    already_voted = db.execute(
        "SELECT 1 FROM votes WHERE voter_id=? AND election_id=?",
        (voter['voter_id'], election_id)
    ).fetchone()
    if already_voted:
        db.close()
        return jsonify(error='You have already voted in this election', code='ALREADY_VOTED'), 409

    # Verify candidate belongs to this election
    valid_candidate = db.execute(
        "SELECT candidate_name FROM candidates WHERE candidate_id=? AND election_id=?",
        (candidate_id, election_id)
    ).fetchone()
    if not valid_candidate:
        db.close()
        return jsonify(error='Invalid candidate selection', code='INVALID_CANDIDATE'), 400

    # ── Face verification ──
    if not face_photo or len(face_photo) < 100:
        db.close()
        return jsonify(
            error='Face photo is required for voting. Please allow camera access.',
            code='NO_FACE_PHOTO'
        ), 400

    matched, similarity = verify_face(voter['face_encoding'], face_photo)
    if not matched:
        db.close()
        if similarity is None:
            if not voter['face_encoding']:
                return jsonify(
                    error='Your face profile is still being processed. Please wait 30 seconds and try again. If this persists, contact support.',
                    code='NO_PROFILE_FACE'
                ), 403
            else:
                return jsonify(
                    error='No face detected in the camera. Please ensure your face is clearly visible and well-lit.',
                    code='NO_FACE_DETECTED'
                ), 403
        return jsonify(
            error=f'Face verification failed (similarity: {similarity}). Your face does not match our records.',
            code='FACE_MISMATCH'
        ), 403

    # ── Record vote ──
    # Render puts the real user IP in the X-Forwarded-For header
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip_address:
        ip_address = ip_address.split(',')[0].strip()
        
    location   = fetch_location(ip_address)
    timestamp  = datetime.utcnow().isoformat() + 'Z'

    db.execute(
        "INSERT INTO votes(voter_id, candidate_id, election_id, ip_address, location, timestamp) VALUES(?,?,?,?,?,?)",
        (voter['voter_id'], candidate_id, election_id, ip_address, location, timestamp)
    )
    db.execute(
        "UPDATE voters SET has_voted=1 WHERE voter_id=?",
        (voter['voter_id'],)
    )
    db.commit()
    db.close()

    # Send vote confirmation email in the background
    if voter['email']:
        def _send_vote_email(v_email, v_name, e_title, c_name, v_timestamp):
            with app.app_context():
                try:
                    import time
                    time.sleep(1) # Let HTTP finish
                    from email_service import send_vote_confirmation_email
                    send_vote_confirmation_email(mail, v_email, v_name, e_title, c_name, v_timestamp)
                    logger.info(f"Vote confirmation email sent for voter ID {voter_id}")
                except Exception as e:
                    logger.error(f"Vote email error: {e}")
        
        voter_id = voter['voter_id']
        c_name = valid_candidate['candidate_name']
        import threading
        t = threading.Thread(target=_send_vote_email, args=(voter['email'], voter['name'], election['election_title'], c_name, timestamp), daemon=True)
        t.start()

    return jsonify(success=True, message='Your vote has been cast successfully!'), 200


# ─────────────────────────────────────────────────────────
# COMPLAINT (PUBLIC)
# ─────────────────────────────────────────────────────────
@app.route('/complaint', methods=['GET', 'POST'])
def complaint():
    if request.method == 'POST':
        name        = request.form.get('name', '').strip()
        email       = request.form.get('email', '').strip()
        roll_number = request.form.get('roll_number', '').strip()
        description = request.form.get('description', '').strip()
        id_card     = request.files.get('id_card')

        errors = []
        if not name or len(name) > 100:
            errors.append('Name is required (max 100 characters).')
        if not email or '@' not in email:
            errors.append('A valid email address is required.')
        if not roll_number or len(roll_number) > 20:
            errors.append('Roll number is required.')
        if not description or len(description) < 20 or len(description) > 1000:
            errors.append('Description must be 20–1000 characters.')

        if errors:
            for e in errors:
                flash(e, 'error')
            return redirect(url_for('complaint'))

        id_card_path = None
        if id_card and id_card.filename:
            ext = os.path.splitext(secure_filename(id_card.filename))[1].lower()
            if ext not in ALLOWED_DOC_EXTS:
                flash('Only JPG, PNG, PDF files allowed.', 'error')
                return redirect(url_for('complaint'))
            fname = f"{uuid.uuid4().hex}_{secure_filename(id_card.filename)}"
            save_path = os.path.join(COMPLAINT_UPLOAD_DIR, fname)
            id_card.save(save_path)
            id_card_path = f"uploads/complaints/{fname}"

        db = get_db()
        cur = db.execute(
            "INSERT INTO complaints(name, email, roll_number, description, id_card_path) VALUES(?,?,?,?,?)",
            (name, email, roll_number, description, id_card_path)
        )
        complaint_id = cur.lastrowid
        db.commit()
        db.close()

        flash(f'Your complaint has been submitted. Reference: #{complaint_id}', 'success')
        return redirect(url_for('complaint'))

    return render_template('complaint.html')


# ─────────────────────────────────────────────────────────
# CARE AUTH
# ─────────────────────────────────────────────────────────
@app.route('/care/login', methods=['GET', 'POST'])
def care_login():
    if session.get('care_logged_in'):
        return redirect(url_for('care_dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username == app.config.get('CARE_USER', 'care') and password == app.config.get('CARE_PASS', 'care123'):
            session['care_logged_in'] = True
            flash('Customer Care access granted.', 'success')
            return redirect(url_for('care_dashboard'))
        else:
            flash('Invalid Care credentials.', 'error')
    return render_template('care/login.html')


@app.route('/care/logout')
def care_logout():
    session.pop('care_logged_in', None)
    flash('Logged out of Customer Care.', 'success')
    return redirect(url_for('care_login'))


# ─────────────────────────────────────────────────────────
# ADMIN AUTH
# ─────────────────────────────────────────────────────────
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if username == app.config['ADMIN_USER'] and password == app.config['ADMIN_PASS']:
            session['admin_logged_in'] = True
            session['admin_user']      = 'admin'
            return redirect(url_for('admin_dashboard'))
        flash('Invalid credentials.', 'error')
    return render_template('admin/login.html')


@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))


# ─────────────────────────────────────────────────────────
# ADMIN DASHBOARD
# ─────────────────────────────────────────────────────────
@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    db = get_db()
    total_voters  = db.execute("SELECT COUNT(*) FROM voters").fetchone()[0]
    registered    = db.execute("SELECT COUNT(*) FROM voters WHERE is_registered=1").fetchone()[0]
    votes_cast    = db.execute("SELECT COUNT(*) FROM votes").fetchone()[0]
    pending_compl = db.execute("SELECT COUNT(*) FROM complaints WHERE status='pending'").fetchone()[0]

    recent_votes = db.execute("""
        SELECT v.timestamp, vt.roll_number, vt.name, e.election_title,
               c.candidate_name, v.location, v.ip_address
        FROM votes v
        JOIN voters vt ON v.voter_id=vt.voter_id
        JOIN elections e ON v.election_id=e.election_id
        JOIN candidates c ON v.candidate_id=c.candidate_id
        ORDER BY v.vote_id DESC LIMIT 5
    """).fetchall()

    elections_raw = db.execute("SELECT * FROM elections ORDER BY election_id DESC LIMIT 6").fetchall()
    elections = []
    for e in elections_raw:
        vote_count = db.execute(
            "SELECT COUNT(*) FROM votes WHERE election_id=?", (e['election_id'],)
        ).fetchone()[0]
        candidate_count = db.execute(
            "SELECT COUNT(*) FROM candidates WHERE election_id=?", (e['election_id'],)
        ).fetchone()[0]
        elections.append({
            'election_id':      e['election_id'],
            'election_title':   e['election_title'],
            'start_date':       e['start_date'],
            'end_date':         e['end_date'],
            'status':           election_status(e['start_date'], e['end_date']),
            'vote_count':       vote_count,
            'candidate_count':  candidate_count,
        })

    try:
        activity_log = db.execute(
            "SELECT action, timestamp as performed_at FROM admin_log ORDER BY log_id DESC LIMIT 10"
        ).fetchall()
    except Exception:
        activity_log = []

    db.close()
    return render_template('admin/dashboard.html',
        total_voters=total_voters,
        registered=registered,
        votes_cast=votes_cast,
        pending_compl=pending_compl,
        recent_votes=recent_votes,
        elections=elections,
        activity_log=activity_log,
    )


# ─────────────────────────────────────────────────────────
# ADMIN — SETTINGS
# ─────────────────────────────────────────────────────────
@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    env_path = os.path.join(app.root_path, '.env')
    if request.method == 'POST':
        resend_key  = request.form.get('resend_key', '').strip()
        mail_from   = request.form.get('mail_from', '').strip()
        if resend_key:
            set_key(env_path, 'RESEND_API_KEY', resend_key)
            os.environ['RESEND_API_KEY'] = resend_key
            import email_service
            email_service.RESEND_API_KEY = resend_key
        if mail_from:
            set_key(env_path, 'MAIL_FROM', mail_from)
            os.environ['MAIL_FROM'] = mail_from
            import email_service
            email_service.MAIL_FROM = mail_from
        log_admin_action('Updated Email Settings', 'System')
        flash('Email settings saved successfully.', 'success')
        return redirect(url_for('admin_settings'))
    current_key  = os.environ.get('RESEND_API_KEY', '')
    current_from = os.environ.get('MAIL_FROM', '')
    return render_template('admin/settings.html', resend_key=current_key, mail_from=current_from)


# ─────────────────────────────────────────────────────────
# ADMIN — VOTERS
# ─────────────────────────────────────────────────────────
@app.route('/admin/voters')
@admin_required
def admin_voters():
    q    = request.args.get('q', '').strip()
    dept = request.args.get('dept', 'all').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per  = 50

    db = get_db()
    query  = "SELECT * FROM voters WHERE 1=1"
    params = []

    if q:
        query += " AND (roll_number LIKE ? OR name LIKE ?)"
        params.extend([f'%{q}%', f'%{q}%'])
    if dept and dept != 'all':
        query += " AND department = ?"
        params.append(dept)

    count_query = f"SELECT COUNT(*) FROM ({query})"
    total = db.execute(count_query, params).fetchone()[0]

    query += " ORDER BY department ASC, roll_number ASC LIMIT ? OFFSET ?"
    params.extend([per, (page-1)*per])
    
    voters = db.execute(query, params).fetchall()
    db.close()

    total_pages = max(1, (total + per - 1) // per)
    return render_template('admin/voters.html',
        voters=voters, q=q, dept=dept, page=page,
        total_pages=total_pages, total=total,
        departments=SVPCET_DEPARTMENTS
    )


@app.route('/admin/voters/add', methods=['POST'])
@admin_required
def admin_voter_add():
    roll_number = request.form.get('roll_number', '').strip().upper()
    name        = request.form.get('name', '').strip()
    department  = request.form.get('department', '').strip()

    if not ROLL_NO_REGEX.match(roll_number):
        flash('Invalid roll number format. Expected: 22G01A4321', 'error')
        return redirect(url_for('admin_voters'))

    db = get_db()
    exists = db.execute("SELECT 1 FROM voters WHERE roll_number=?", (roll_number,)).fetchone()
    if exists:
        db.close()
        flash('Roll number already exists.', 'error')
        return redirect(url_for('admin_voters'))

    db.execute(
        "INSERT INTO voters(roll_number, name, department) VALUES(?,?,?)",
        (roll_number, name, department)
    )
    db.commit()
    db.close()
    flash(f'Voter {roll_number} added successfully.', 'success')
    return redirect(url_for('admin_voters'))


@app.route('/admin/voters/upload-csv', methods=['POST'])
@admin_required
def admin_voter_csv():
    f = request.files.get('csv_file')
    if not f:
        flash('No file uploaded.', 'error')
        return redirect(url_for('admin_voters'))

    content   = f.read().decode('utf-8', errors='replace')
    reader    = csv.DictReader(StringIO(content))
    imported  = 0
    duplicates = 0
    invalid   = 0

    db = get_db()
    for row in reader:
        roll = (row.get('roll_number') or '').strip().upper()
        name = (row.get('name')        or '').strip()
        dept = (row.get('department')  or '').strip()

        if not ROLL_NO_REGEX.match(roll) or not name:
            invalid += 1
            continue

        exists = db.execute("SELECT 1 FROM voters WHERE roll_number=?", (roll,)).fetchone()
        if exists:
            duplicates += 1
            continue

        db.execute(
            "INSERT INTO voters(roll_number, name, department) VALUES(?,?,?)",
            (roll, name, dept)
        )
        imported += 1

    db.commit()
    db.close()
    flash(f'Imported {imported} rows. Skipped {duplicates} duplicates. {invalid} invalid.', 'success')
    return redirect(url_for('admin_voters'))


@app.route('/admin/voters/delete/<int:voter_id>', methods=['POST'])
@admin_required
def admin_voter_delete(voter_id):
    db = get_db()
    has_votes = db.execute("SELECT 1 FROM votes WHERE voter_id=?", (voter_id,)).fetchone()
    if has_votes:
        db.close()
        flash('Cannot delete voter with existing votes.', 'error')
        return redirect(url_for('admin_voters'))
    db.execute("DELETE FROM voters WHERE voter_id=?", (voter_id,))
    db.commit()
    db.close()
    flash('Voter deleted.', 'success')
    return redirect(url_for('admin_voters'))


# ─────────────────────────────────────────────────────────
# ADMIN — ELECTIONS
# ─────────────────────────────────────────────────────────
@app.route('/admin/elections')
@admin_required
def admin_elections():
    db = get_db()
    elections_raw = db.execute("SELECT * FROM elections ORDER BY election_id DESC").fetchall()
    elections = []
    for e in elections_raw:
        # Get eligible departments or roll count
        if (e['eligible_type'] or 'department') == 'department':
            elig_info = ', '.join([r['department'] for r in db.execute(
                "SELECT department FROM election_eligible_depts WHERE election_id=?",
                (e['election_id'],)
            ).fetchall()]) or '—'
        else:
            count = db.execute(
                "SELECT COUNT(*) FROM election_eligible_rolls WHERE election_id=?",
                (e['election_id'],)
            ).fetchone()[0]
            elig_info = f"{count} specific roll no(s)"

        vote_count = db.execute(
            "SELECT COUNT(*) FROM votes WHERE election_id=?", (e['election_id'],)
        ).fetchone()[0]

        elections.append({
            'election_id':       e['election_id'],
            'election_title':    e['election_title'],
            'description':       e['description'] or '',
            'position_role':     e['position_role'] or '',
            'eligible_type':     e['eligible_type'] or 'department',
            'elig_info':         elig_info,
            'start_date':        e['start_date'],
            'end_date':          e['end_date'],
            'status':            election_status(e['start_date'], e['end_date']),
            'vote_count':        vote_count,
            'results_published': e['results_published'],
        })
    db.close()
    return render_template('admin/elections.html', elections=elections, departments=SVPCET_DEPARTMENTS)


@app.route('/admin/elections/create', methods=['GET', 'POST'])
@admin_required
def admin_election_create():
    if request.method == 'GET':
        return render_template('admin/election_create.html', departments=SVPCET_DEPARTMENTS)

    title         = request.form.get('election_title', '').strip()
    description   = request.form.get('description', '').strip()
    position_role = request.form.get('position_role', '').strip()
    start_date    = request.form.get('start_date', '').strip()
    end_date      = request.form.get('end_date', '').strip()
    eligible_type = request.form.get('eligible_type', 'department').strip()

    if not title:
        flash('Election title is required.', 'error')
        return redirect(url_for('admin_elections'))

    try:
        s = datetime.strptime(start_date, '%Y-%m-%d').date()
        e = datetime.strptime(end_date,   '%Y-%m-%d').date()
        if s >= e:
            flash('Start date must be before end date.', 'error')
            return redirect(url_for('admin_election_create'))
    except ValueError:
        flash('Invalid date format.', 'error')
        return redirect(url_for('admin_election_create'))

    db = get_db()
    cur = db.execute(
        "INSERT INTO elections(election_title, description, position_role, eligible_type, start_date, end_date, roll_start, roll_end) VALUES(?,?,?,?,?,?, '', '')",
        (title, description, position_role, eligible_type, start_date, end_date)
    )
    election_id = cur.lastrowid

    # ── Handle eligibility ──
    if eligible_type == 'department':
        selected_depts = request.form.getlist('departments[]')
        for dept in selected_depts:
            if dept:
                db.execute(
                    "INSERT INTO election_eligible_depts(election_id, department) VALUES(?,?)",
                    (election_id, dept)
                )

    else:  # specific_rolls
        roll_numbers = set()

        # From textarea
        rolls_text = request.form.get('rolls_text', '').strip()
        if rolls_text:
            for line in rolls_text.splitlines():
                val = line.strip().upper()
                if ROLL_NO_REGEX.match(val):
                    roll_numbers.add(val)

        # From Excel upload
        excel_file = request.files.get('rolls_excel')
        if excel_file and excel_file.filename:
            ext = os.path.splitext(secure_filename(excel_file.filename))[1].lower()
            if ext in ALLOWED_EXCEL_EXTS:
                parsed = parse_rolls_from_excel(BytesIO(excel_file.read()))
                roll_numbers.update(parsed)

        for roll in roll_numbers:
            db.execute(
                "INSERT INTO election_eligible_rolls(election_id, roll_number) VALUES(?,?)",
                (election_id, roll)
            )

    # ── Handle Candidates ──
    idx = 0
    while True:
        c_name = request.form.get(f'candidate_name_{idx}')
        if c_name is None:
            break
        
        c_name = c_name.strip()
        if c_name:
            c_pos   = request.form.get(f'candidate_position_{idx}', '').strip()
            c_party = request.form.get(f'candidate_party_{idx}', '').strip()
            c_desc  = request.form.get(f'candidate_desc_{idx}', '').strip()
            c_sym   = request.files.get(f'candidate_symbol_{idx}')
            
            sym_path = None
            if c_sym and c_sym.filename:
                ext = os.path.splitext(secure_filename(c_sym.filename))[1].lower()
                if ext in ALLOWED_IMAGE_EXTS:
                    fname = f"{uuid.uuid4().hex}_{secure_filename(c_sym.filename)}"
                    c_sym.save(os.path.join(SYMBOL_UPLOAD_DIR, fname))
                    sym_path = f"uploads/symbols/{fname}"
            
            db.execute(
                "INSERT INTO candidates(election_id, candidate_name, description, position, party_name, symbol_path) VALUES(?,?,?,?,?,?)",
                (election_id, c_name, c_desc, c_pos, c_party, sym_path)
            )
        idx += 1

    db.commit()

    # ── Send announcement emails to eligible registered voters ──
    recipients = get_eligible_voter_emails(db, election_id)
    db.close()

    log_admin_action('Created Election', title)

    def send_announcements():
        with app.app_context():
            for email_addr, voter_name in recipients:
                try:
                    send_election_announcement(
                        mail, email_addr, voter_name,
                        title, description, position_role, start_date, end_date
                    )
                except Exception as ex:
                    logger.error(f"Announcement email failed to {email_addr}: {ex}")

    if recipients:
        t = threading.Thread(target=send_announcements, daemon=True)
        t.start()

    flash(f'Election "{title}" created. Announcement sent to {len(recipients)} eligible voter(s).', 'success')
    return redirect(url_for('admin_elections'))


@app.route('/admin/elections/<int:election_id>/delete', methods=['POST'])
@admin_required
def admin_election_delete(election_id):
    db = get_db()
    has_votes = db.execute("SELECT 1 FROM votes WHERE election_id=?", (election_id,)).fetchone()
    if has_votes:
        db.close()
        flash('Cannot delete an election that has votes recorded.', 'error')
        return redirect(url_for('admin_elections'))
    db.execute("DELETE FROM candidates WHERE election_id=?", (election_id,))
    db.execute("DELETE FROM election_eligible_depts WHERE election_id=?", (election_id,))
    db.execute("DELETE FROM election_eligible_rolls WHERE election_id=?", (election_id,))
    db.execute("DELETE FROM elections WHERE election_id=?", (election_id,))
    db.commit()
    db.close()
    log_admin_action('Deleted Election', str(election_id))
    flash('Election deleted.', 'success')
    return redirect(url_for('admin_elections'))


@app.route('/admin/elections/<int:election_id>/results')
@admin_required
def election_results(election_id):
    db = get_db()
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()
    if not election:
        db.close()
        abort(404)

    candidates = db.execute("""
        SELECT c.candidate_id, c.candidate_name, c.position, c.party_name, c.symbol_path,
               COUNT(v.vote_id) as vote_count
        FROM candidates c
        LEFT JOIN votes v ON c.candidate_id=v.candidate_id AND v.election_id=c.election_id
        WHERE c.election_id=?
        GROUP BY c.candidate_id
        ORDER BY vote_count DESC
    """, (election_id,)).fetchall()

    total_votes = db.execute(
        "SELECT COUNT(*) FROM votes WHERE election_id=?", (election_id,)
    ).fetchone()[0]

    max_votes = candidates[0]['vote_count'] if candidates else 0

    db.close()

    candidates_list = []
    for c in candidates:
        bar_pct  = round((c['vote_count'] / max_votes * 100), 1) if max_votes > 0 else 0
        vote_pct = round((c['vote_count'] / total_votes * 100), 1) if total_votes > 0 else 0
        candidates_list.append({
            'candidate_id':   c['candidate_id'],
            'candidate_name': c['candidate_name'],
            'position':       c['position'],
            'party_name':     c['party_name'],
            'symbol_path':    c['symbol_path'],
            'vote_count':     c['vote_count'],
            'bar_pct':        bar_pct,
            'vote_pct':       vote_pct,
        })

    return render_template('admin/election_results.html',
        election=election,
        candidates=candidates_list,
        total_votes=total_votes,
        turnout_pct=0,
        status=election_status(election['start_date'], election['end_date'])
    )


@app.route('/admin/elections/<int:election_id>/publish-results', methods=['POST'])
@admin_required
def admin_publish_results(election_id):
    db = get_db()
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()
    if not election:
        db.close()
        abort(404)

    # Find winner
    winner = db.execute("""
        SELECT c.candidate_name, c.party_name, COUNT(v.vote_id) as vote_count
        FROM candidates c
        LEFT JOIN votes v ON c.candidate_id=v.candidate_id AND v.election_id=c.election_id
        WHERE c.election_id=?
        GROUP BY c.candidate_id
        ORDER BY vote_count DESC
        LIMIT 1
    """, (election_id,)).fetchone()

    total_votes = db.execute(
        "SELECT COUNT(*) FROM votes WHERE election_id=?", (election_id,)
    ).fetchone()[0]

    db.execute(
        "UPDATE elections SET results_published=1 WHERE election_id=?", (election_id,)
    )
    db.commit()

    # Send results emails
    recipients = get_eligible_voter_emails(db, election_id)
    db.close()

    log_admin_action('Published Results', election['election_title'])

    if winner:
        def send_results_emails():
            with app.app_context():
                for email_addr, voter_name in recipients:
                    try:
                        send_election_results(
                            mail, email_addr, voter_name,
                            election['election_title'],
                            election['position_role'] or '',
                            winner['candidate_name'],
                            winner['party_name'] or '',
                            total_votes
                        )
                    except Exception as ex:
                        logger.error(f"Results email failed to {email_addr}: {ex}")

        t = threading.Thread(target=send_results_emails, daemon=True)
        t.start()
        flash(f'Results published. Winner: {winner["candidate_name"]}. Emails sent to {len(recipients)} voter(s).', 'success')
    else:
        flash('Results published. No votes were cast.', 'warning')

    return redirect(url_for('election_results', election_id=election_id))


# ─────────────────────────────────────────────────────────
# ADMIN — LIVE VOTE TRACKER
# ─────────────────────────────────────────────────────────
@app.route('/admin/elections/<int:election_id>/live')
@admin_required
def election_live(election_id):
    db = get_db()
    election = db.execute(
        "SELECT * FROM elections WHERE election_id=?", (election_id,)
    ).fetchone()
    if not election:
        db.close()
        abort(404)

    vote_rows = db.execute("""
        SELECT
            v.vote_id,
            v.timestamp,
            vt.roll_number,
            vt.name     AS voter_name,
            vt.department,
            c.candidate_name,
            c.party_name,
            c.position  AS candidate_position,
            v.ip_address,
            v.location
        FROM votes v
        JOIN voters    vt ON v.voter_id    = vt.voter_id
        JOIN candidates c ON v.candidate_id = c.candidate_id
        WHERE v.election_id = ?
        ORDER BY v.vote_id DESC
    """, (election_id,)).fetchall()

    total_votes = len(vote_rows)
    eligible_count = 0
    etype = election['eligible_type'] or 'department'
    if etype == 'all':
        eligible_count = db.execute(
            "SELECT COUNT(*) FROM voters WHERE is_registered=1"
        ).fetchone()[0]
    elif etype == 'department':
        depts = [r['department'] for r in db.execute(
            "SELECT department FROM election_eligible_depts WHERE election_id=?", (election_id,)
        ).fetchall()]
        if depts:
            ph = ','.join('?' * len(depts))
            eligible_count = db.execute(
                f"SELECT COUNT(*) FROM voters WHERE department IN ({ph}) AND is_registered=1", depts
            ).fetchone()[0]
    else:
        eligible_count = db.execute(
            "SELECT COUNT(*) FROM election_eligible_rolls WHERE election_id=?", (election_id,)
        ).fetchone()[0]

    db.close()

    votes = [dict(row) for row in vote_rows]
    turnout = round((total_votes / eligible_count * 100), 1) if eligible_count > 0 else 0

    return render_template('admin/election_live.html',
        election=election,
        votes=votes,
        total_votes=total_votes,
        eligible_count=eligible_count,
        turnout=turnout,
        status=election_status(election['start_date'], election['end_date'])
    )


@app.route('/admin/elections/<int:election_id>/live/data')
@admin_required
def election_live_data(election_id):
    """JSON endpoint for AJAX polling in live tracker."""
    db = get_db()
    vote_rows = db.execute("""
        SELECT
            v.vote_id,
            v.timestamp,
            vt.roll_number,
            vt.name     AS voter_name,
            vt.department,
            c.candidate_name,
            c.party_name,
            v.ip_address,
            v.location
        FROM votes v
        JOIN voters     vt ON v.voter_id     = vt.voter_id
        JOIN candidates c  ON v.candidate_id = c.candidate_id
        WHERE v.election_id = ?
        ORDER BY v.vote_id DESC
    """, (election_id,)).fetchall()
    db.close()
    return jsonify(votes=[dict(r) for r in vote_rows], total=len(vote_rows))


# ─────────────────────────────────────────────────────────
# ADMIN — CANDIDATES
# ─────────────────────────────────────────────────────────
@app.route('/admin/candidates')
@admin_required
def admin_candidates():
    db = get_db()
    elections  = db.execute("SELECT * FROM elections ORDER BY election_id DESC").fetchall()
    candidates = db.execute("""
        SELECT c.*, e.election_title
        FROM candidates c
        JOIN elections e ON c.election_id=e.election_id
        ORDER BY c.election_id DESC, c.candidate_id ASC
    """).fetchall()
    db.close()
    return render_template('admin/candidates.html', elections=elections, candidates=candidates)


@app.route('/admin/candidates/add', methods=['POST'])
@admin_required
def admin_candidate_add():
    election_id    = request.form.get('election_id')
    candidate_name = request.form.get('candidate_name', '').strip()
    position       = request.form.get('position', '').strip()
    description    = request.form.get('description', '').strip()
    party_name     = request.form.get('party_name', '').strip()
    symbol         = request.files.get('symbol')

    if not election_id or not candidate_name or not position:
        flash('Election, name, and position are required.', 'error')
        return redirect(url_for('admin_candidates'))

    symbol_path = None
    if symbol and symbol.filename:
        ext = os.path.splitext(secure_filename(symbol.filename))[1].lower()
        if ext not in ALLOWED_IMAGE_EXTS:
            flash('Symbol must be an image file.', 'error')
            return redirect(url_for('admin_candidates'))
        fname = f"{uuid.uuid4().hex}_{secure_filename(symbol.filename)}"
        symbol.save(os.path.join(SYMBOL_UPLOAD_DIR, fname))
        symbol_path = f"uploads/symbols/{fname}"

    db = get_db()
    db.execute(
        "INSERT INTO candidates(election_id, candidate_name, description, position, party_name, symbol_path) VALUES(?,?,?,?,?,?)",
        (election_id, candidate_name, description, position, party_name, symbol_path)
    )
    db.commit()
    db.close()
    flash('Candidate added successfully.', 'success')
    return redirect(url_for('admin_candidates'))


@app.route('/admin/candidates/delete/<int:candidate_id>', methods=['POST'])
@admin_required
def admin_candidate_delete(candidate_id):
    db = get_db()
    has_votes = db.execute("SELECT 1 FROM votes WHERE candidate_id=?", (candidate_id,)).fetchone()
    if has_votes:
        db.close()
        flash('Cannot delete candidate with existing votes.', 'error')
        return redirect(url_for('admin_candidates'))
    db.execute("DELETE FROM candidates WHERE candidate_id=?", (candidate_id,))
    db.commit()
    db.close()
    flash('Candidate deleted.', 'success')
    return redirect(url_for('admin_candidates'))


# ─────────────────────────────────────────────────────────
# CARE — DASHBOARD & COMPLAINTS
# ─────────────────────────────────────────────────────────
@app.route('/care/dashboard')
@care_required
def care_dashboard():
    status_filter = request.args.get('status', 'all')
    db = get_db()
    if status_filter in ('pending', 'resolved', 'rejected'):
        complaints = db.execute(
            "SELECT * FROM complaints WHERE status=? ORDER BY complaint_id DESC",
            (status_filter,)
        ).fetchall()
    else:
        complaints = db.execute(
            "SELECT * FROM complaints ORDER BY complaint_id DESC"
        ).fetchall()
    db.close()
    return render_template('care/dashboard.html',
        complaints=complaints, title='Customer Care Dashboard', status_filter=status_filter
    )


@app.route('/care/complaints/<int:complaint_id>')
@care_required
def care_complaint_detail(complaint_id):
    db = get_db()
    complaint = db.execute(
        "SELECT * FROM complaints WHERE complaint_id=?", (complaint_id,)
    ).fetchone()
    if not complaint:
        db.close()
        abort(404)
    voter = db.execute(
        "SELECT * FROM voters WHERE roll_number=?", (complaint['roll_number'],)
    ).fetchone()
    db.close()
    return render_template('care/complaint_detail.html', complaint=complaint, voter=voter)


@app.route('/care/complaints/<int:complaint_id>/resolve', methods=['POST'])
@care_required
def care_complaint_resolve(complaint_id):
    db = get_db()
    complaint = db.execute("SELECT * FROM complaints WHERE complaint_id=?", (complaint_id,)).fetchone()
    if not complaint:
        db.close()
        abort(404)

    admin_response = request.form.get('admin_response', '').strip()
    new_name  = request.form.get('new_name',  '').strip()
    new_email = request.form.get('new_email', '').strip()
    new_phone = request.form.get('new_phone', '').strip()

    if new_name or new_email or new_phone:
        fields, vals = [], []
        if new_name:  fields.append('name=?');  vals.append(new_name)
        if new_email: fields.append('email=?'); vals.append(new_email)
        if new_phone: fields.append('phone=?'); vals.append(new_phone)
        vals.append(complaint['roll_number'])
        db.execute(f"UPDATE voters SET {', '.join(fields)} WHERE roll_number=?", vals)

    db.execute("UPDATE complaints SET status='resolved', admin_response=? WHERE complaint_id=?",
               (admin_response, complaint_id))
    db.commit()
    db.close()

    if admin_response and complaint['email']:
        def _send_care_resolve_email(c_email, c_name, c_id, a_resp):
            send_care_response_email(c_email, c_name, c_id, 'resolved', a_resp)

        t = threading.Thread(target=_send_care_resolve_email, args=(complaint['email'], complaint['name'], complaint_id, admin_response), daemon=True)
        t.start()

    flash('Support case resolved and voter notified.', 'success')
    return redirect(url_for('care_dashboard'))


@app.route('/care/complaints/<int:complaint_id>/reject', methods=['POST'])
@care_required
def care_complaint_reject(complaint_id):
    db = get_db()
    complaint = db.execute("SELECT * FROM complaints WHERE complaint_id=?", (complaint_id,)).fetchone()
    if not complaint:
        db.close()
        abort(404)

    admin_response = request.form.get('admin_response', '').strip()
    db.execute("UPDATE complaints SET status='rejected', admin_response=? WHERE complaint_id=?",
               (admin_response, complaint_id))
    db.commit()
    db.close()

    if admin_response and complaint['email']:
        def _send_care_reject_email(c_email, c_name, c_id, a_resp):
            send_care_response_email(c_email, c_name, c_id, 'rejected', a_resp)

        t = threading.Thread(target=_send_care_reject_email, args=(complaint['email'], complaint['name'], complaint_id, admin_response), daemon=True)
        t.start()

        t = threading.Thread(target=_send_care_reject_email, args=(complaint['email'], complaint['name'], complaint_id, admin_response), daemon=True)
        t.start()

    flash('Support case rejected.', 'success')
    return redirect(url_for('care_dashboard'))


# ─────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────
with app.app_context():
    init_db()

# NOTE: ArcFace model is intentionally NOT preloaded here.
# Preloading in the gunicorn master process causes the model to be
# duplicated into every forked worker (fork = copy-on-write), using 2x RAM
# and causing OOM crashes on Render's 512MB free tier.
# The model lazy-loads on the first background encoding thread run.

if __name__ == '__main__':
    # In local dev, preload for convenience (no forking happens)
    from face_service import get_face_app
    get_face_app()
    app.run(debug=True, port=5000, threaded=True)

